# -*- coding:utf8 -*-
'''contains basic class
   last modified April 3, 2018'''

import os
from openpyxl import *
# import openpyxl
from utility import *
import logging
import yaml


# global logger
# def set_log(path):
#    import logging
#    import time
#    global logger
#    logger = logging.getLogger()
#    logger.setLevel(logging.DEBUG)
#    formatter_file = logging.Formatter(
#        '%(asctime)s - %(levelname)s: - %(message)s',
#        datefmt='%Y-%m-%d %H:%M:%S')
#    formatter_screen = logging.Formatter(
#        '%(levelname)s: - %(message)s',
#        datefmt='%Y-%m-%d %H:%M:%S')
#    # output to file
#    log_path = ref_dict['Path']+'auto_test'+ \
#               time.strftime('%Y%m%d%H%M%S', \
#               time.localtime(time.time()))+'.log',
#    fl = logging.FileHandler(log_path)
#    fl.setLevel(logging.DEBUG)
#    fl.setFormatter(formatter_file)
#    # outpur to screen
#    sr = logging.StreamHandler()
#    sr.setLevel(logging.WARNING)
#    sr.setFormatter(formatter_screen)
#    # add Handler of file and screen
#    logger.addHandler(fl)
#    logger.addHandler(sr)
#    return logger


def load_file(filename):
    '''read input file'''
    import yaml
    import time
    file = open(filename)
    data = yaml.safe_load(file)
    folder_list = []
    ref_dict = data['ReferenceFolder']
    ref_folder = Folder(ref_dict['Name'], ref_dict['Path'])
    logging.basicConfig(level=logging.INFO,
                        format='%(levelname)s %(message)s',
                        filename=ref_dict['Path'] + '\\auto_test' + \
                                 time.strftime('%Y%m%d%H%M%S', \
                                               time.localtime(time.time())) + '.log',
                        filemode='w')
    ref_folder.get_files()
    logging.info('succeed to scan ' + ref_dict['Path'])
    for folder_dict in data['FolderList']:
        folder = Folder(folder_dict['Name'], folder_dict['Path'])
        folder.get_files()
        logging.info('succeed to scan ' + folder_dict['Path'])
        folder_list.append(folder)
    return ref_folder, folder_list


def compare_folders(folder_A, folder_B):
    '''compare folder_A and folder_B,
       folder_A as reference, res_folder: result folder'''
    import os
    res_folder = Folder('compare_result', os.path.join(folder_B.path, 'compare_result'))
    for ifile in folder_A.files:
        xfile = search_folders(folder_B, ifile)
        if xfile is None:
            logging.critical('file ' + ifile.fname + ' not found in ' + folder_B.path)
            continue
        print('processing' + ifile.fname)
        xfile.compare_files(ifile, res_folder.path)
    return None


def compare_folders1(pre_list, pro_list, file_path):

    import os
    res_folder = Folder('compare_result', os.path.join(file_path, 'compare_result'))
    for pre_file in pre_list:
        pre_file_name = pre_file.split('\\')[1]
        for pro_file in pro_list:
            pro_file_name = pro_file.split('\\')[1]
            if pre_file_name == pro_file_name:
                pro_file.compare_files(pre_file, res_folder.path)
    return None



def list_files(path, base_path, list):
    '''return list of all files and ziped files in base_path,
       including sub-folders '''
    # extract_zip(path)
    file_list = os.listdir(path)
    for ifile in file_list:
        file_name, file_ext = os.path.splitext(ifile)
        if file_ext == '.xlsx':
            file_obj = Xls(path, base_path, ifile, file_ext)
            list.append(file_obj)
        elif file_ext == '.csv':
            file_obj = Csv(path, base_path, ifile, file_ext)
            list.append(file_obj)
        elif file_ext == '.txt':
            file_obj = Txt(path, base_path, ifile, file_ext)
            list.append(file_obj)
        elif file_ext == '' and file_name.lower() != 'summary':
            # recursion on this folder
            list_in = []
            list.extend(list_files(os.path.join(path, ifile), base_path, list_in))
    return list


class Folder():
    def __init__(self, fname, fpath):
        self.name = fname
        self.path = fpath
        self.files = []

    def get_files(self):
        base_path = self.path
        list = []
        self.files = list_files(self.path, base_path, list)


class File():
    def __init__(self, path, base_path, file, file_ext):
        self.fname = os.path.join(path.replace(base_path, ""), file)
        self.ext = file_ext
        self.path = os.path.join(path, file)

    def compare_files(self, file, path):
        pass


class Csv(File):
    def __init__(self, path, base_path, file, file_ext):
        File.__init__(self, path, base_path, file, file_ext)

    def compare_files(self, file, path):
        import csv
        import os
        is_match = True
        res_file_path = os.path.split(path + self.fname)[0]
        create_folder(res_file_path)
        file_out = open(path + self.fname, 'w', newline='')
        file_write = csv.writer(file_out)
        # load data
        ref_data = open(file.path).readlines()
        tar_data = open(self.path).readlines()
        if len(ref_data) != len(tar_data):
            logging.critical('Number of rows of ' + file.path + \
                             ' and ' + self.path + ' not match')
            is_match = False
        for iline in range(min(len(ref_data), len(tar_data))):
            # 'normalize' line content
            ref_line = text_to_list(ref_data[iline])
            tar_line = text_to_list(tar_data[iline])
            res_line = []
            if len(ref_line) != len(tar_line):
                logging.critical('Number of data in row ' + str(iline + 1) + ' of ' + file.path + \
                                 ' and ' + self.path + ' not match')
                is_match = False
            for idata in range(min(len(ref_line), len(tar_line))):
                # calculate difference
                ref_value = ref_line[idata].strip()
                tar_value = tar_line[idata].strip()
                if is_value(ref_value) != is_value(tar_value):
                    logging.critical('Data type of columns of ' + file.path + \
                                     ' and ' + self.path + ' not match in cell (' \
                                     + str(iline + 1) + ',' + str(idata + 1) + ')')
                    is_match = False
                if (is_value(ref_value) and is_value(tar_value)):
                    value = cal_difference(ref_value, tar_value)
                    if value != '0.0000%':
                        is_match = False
                elif ref_value != tar_value:
                    value = 'Diff!'
                    is_match = False
                else:
                    value = ref_value
                res_line.append(value)
            # write line to result
            file_write.writerow(res_line)
        file_out.close()
        print('%s%s%s' % ('file ', self.fname, ' compare complete'))
        if is_match:
            logging.info('%s%s%s' % ('file ', self.path, ' compare complete, perferctly match!'))
        else:
            logging.error('%s%s%s' % ('file ', self.path, ' compare complete, somewhere different.'))
        return None


class Xls(File):
    def __init__(self, path, base_path, file, file_ext):
        File.__init__(self, path, base_path, file, file_ext)

    def compare_files(self, file, path):
        import os
        from openpyxl.utils import get_column_letter
        # load excel and content sheet
        ref_data = load_workbook(file.path)
        tar_data = load_workbook(self.path)
        excel_out = Workbook()
        is_match = True
        iter = 0
        for isheet in ref_data.sheetnames:
            ref_sheet = ref_data[isheet]
            try:
                tar_sheet = tar_data[isheet]
            except:
                logging.critical('sheet ' + isheet + ' not exists in ' + self.path)
                return None
            # create new excel sheet
            if not iter:
                # active the initial sheet
                excel_sheet = excel_out.active
            else:
                # add sheets afterward
                excel_sheet = excel_out.create_sheet()
            excel_sheet.title = isheet + '_diff'
            # compare row number
            ref_nrow = ref_sheet.max_row
            tar_nrow = tar_sheet.max_row
            if ref_nrow != tar_nrow:
                logging.critical('Number of rows of ' + file.path + \
                                 ' and ' + self.path + ' not match')
                is_match = False
            # compare column number
            ref_ncol = ref_sheet.max_column
            tar_ncol = tar_sheet.max_column
            if ref_ncol != tar_ncol:
                logging.critical('Number of columns of ' + file.path + \
                                 ' and ' + self.path + ' not match')
                is_match = False
            # compare cell by cell
            for irow in range(ref_nrow):
                for icol in range(ref_ncol):
                    # get value from cell
                    col_letter = get_column_letter(icol + 1)
                    ref_value = ref_sheet['%s%s' % (col_letter, irow + 1)].value
                    tar_value = tar_sheet['%s%s' % (col_letter, irow + 1)].value
                    # calculate difference and write to result
                    if is_value(ref_value) != is_value(tar_value):
                        logging.critical('Data type of columns of ' + \
                                         file.path + ' and ' + self.path + \
                                         ' not match in cell [' + col_letter + str(irow + 1) + ']')
                        value = 'Diff!'
                        is_match = False
                    elif (is_value(ref_value) and is_value(tar_value)):
                        value = cal_difference(ref_value, tar_value)
                        if value != '0.0000%':
                            is_match = False
                    elif ref_value != tar_value:
                        value = 'Diff!'
                        is_match = False
                    else:
                        value = ref_sheet['%s%s' % (col_letter, irow + 1)].value
                    excel_sheet['%s%s' % (col_letter, irow + 1)] = value
            iter += 1
        # save workbook
        res_file_path = os.path.split(path + self.fname)[0]
        create_folder(res_file_path)
        excel_out.save(path + self.fname)
        print('%s%s%s' % ('file ', self.fname, ' compare complete'))
        if is_match:
            logging.info('%s%s%s' % ('file ', self.path, ' compare complete, perferctly match!'))
        else:
            logging.error('%s%s%s' % ('file ', self.path, ' compare complete, somewhere different.'))
        return None


class Txt(File):
    def __init__(self, path, base_path, file, file_ext):
        File.__init__(self, path, base_path, file, file_ext)

    def compare_files(self, file, path):
        import os
        is_match = True
        # load data
        ref_data = open(file.path).readlines()
        tar_data = open(self.path).readlines()
        if len(ref_data) != len(tar_data):
            logging.critical('Number of rows of ' + file.path + \
                             ' and ' + self.path + ' not match')
            is_match = False
        res_file_path = os.path.split(path + self.fname)[0]
        # create result file
        create_folder(res_file_path)
        file_out = open(path + self.fname, 'w')
        for iline in range(min(len(ref_data), len(tar_data))):
            # 'normalize' line content
            ref_line = text_to_list(ref_data[iline])
            tar_line = text_to_list(tar_data[iline])
            res_line = ''
            if len(ref_line) != len(tar_line):
                logging.critical('Number of data in row ' + str(iline + 1) + ' of ' + file.path + \
                                 ' and ' + self.path + ' not match')
                is_match = False
            for idata in range(min(len(ref_line), len(tar_line))):
                ref_value = ref_line[idata].strip()
                tar_value = tar_line[idata].strip()
                if is_value(ref_value) != is_value(tar_value):
                    logging.critical('Data type of columns of ' + file.path +
                                     ' and ' + self.path + ' not match in cell ('
                                     + str(iline + 1) + ',' + str(idata + 1) + ')')
                    is_match = False
                # calculate difference
                if (is_value(ref_value) and is_value(tar_value)):
                    value = cal_difference(ref_value, tar_value)
                    if value != '0.0000%':
                        is_match = False
                elif ref_value != tar_value:
                    value = 'Diff!'
                    is_match = False
                else:
                    value = ref_value
                res_line = res_line + '\t' + value
            # write line to result
            file_out.write(res_line + '\n')
        file_out.close()
        print('%s%s%s' % ('file ', self.path, ' compare complete'))
        if is_match:
            logging.info('%s%s%s' % ('file ', self.path, ' compare complete, perferctly match!'))
        else:
            logging.error('%s%s%s' % ('file ', self.path, ' compare complete, somewhere different.'))
        return None


def get_all_xct_filepath(file_path):
    path_names = get_all_filepath(file_path)
    for file_zip in path_names:
        dst_dir = file_zip.split('\\')[0]
        unzip_file(file_zip, dst_dir)
    # 再次获取解包后的所有文件路径，过滤xlsx csv txt类型文件
    path_names1 = get_all_filepath(file_path)
    xct_list = get_xct_filepath(path_names1)
    return xct_list


def compare_result(file_path, file_name, exp_value):
    import os
    from openpyxl.utils import get_column_letter

    # load excel and content sheet
    file_ext = os.path.splitext(file_name)[1].strip('.')
    ref_data = load_workbook(file_path + file_name)
    ref_sheet = ref_data.sheetnames[0]
    ref_nrow = ref_sheet.max_row
    ref_ncol = ref_sheet.max_column
    error_count = 0
    for irow in range(ref_nrow):
        for icol in range(ref_ncol):
            col_letter = get_column_letter(icol + 1)
            ref_value = ref_sheet['%s%s' % (col_letter, irow + 1)].value
            if is_value(ref_value) >= exp_value:
                error_count += 1
    print(file_name, ' compare complete')
    if error_count > 0:
        logging.info(file_name + 'has' + str(error_count) + 'errors,please check file')
    else:
        logging.info(file_name + 'compare passed')
    return None
