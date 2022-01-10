#!/usr/bin/env python
# -*- encoding: utf-8 -*-
"""
@file : check_compare.py
@comment:
@date : 2021/12/14 15:03
@author : fei.li
@version : 1.0
"""

from openpyxl import *
import logging
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import yaml
import csv


def compare_xls_result(file_name, exp_value):
    # load excel and content sheet
    ref_data = load_workbook(filename=file_name)
    ref_sheet = ref_data[ref_data.sheetnames[0]]
    ref_nrow = ref_sheet.max_row
    ref_ncol = ref_sheet.max_column
    error_count = 0
    for irow in range(ref_nrow):
        for icol in range(ref_ncol):
            col_letter = get_column_letter(icol + 1)
            ref_value = ref_sheet['%s%s' % (col_letter, irow + 1)].value
            ref_value1 = to_float(ref_value)
            if ref_value1 is not None and ref_value1 >= float(exp_value):
                error_count += 1
    print(file_name, ' compare completed')
    if error_count > 0:
        logging.info(file_name + 'has' + str(error_count) + 'errors,please check file')
        return file_name, error_count
    else:
        logging.info(file_name + 'compare passed')
        return None


def compare_txt_result(file_name, exp_value):
    file = file_name
    exp_value = float(exp_value)
    error_count = 0
    txt_list = []
    with open(file, 'r') as fp:
        for line in fp.readlines():
            txt_list.append(line.strip())
    txt_list1 = []
    for i in txt_list[1:]:
        a = i.split('\t')
        txt_list1.append(a[0])
        txt_list1.append(a[1])
    for i in txt_list1:
        if to_float(i) >= exp_value:
            error_count += 1
    if error_count > 0:
        logging.info(file_name + 'has' + str(error_count) + 'errors,please check file')
        return file_name, error_count
    else:
        logging.info(file_name + 'compare passed')
        return None


def compare_csv_result(file_path, exp_value):
    csv_read = csv.reader(open(file_path))
    error_count = 0
    for row in csv_read:
        # print(row)
        if not row:
            continue
        elif row[0] == 'ACE Version:':
            break
        else:
            # print(row)
            for i in row:
                if len(i) > 3 and i[-1] == '%':
                    j = to_float(i)
                    if j > float(exp_value):
                        error_count += 1
    if error_count > 0:
        logging.info(file_path + 'has' + str(error_count) + 'errors, please check file')
        return file_path, error_count
    else:
        logging.info(file_path + 'compare passed')
        return None


def to_float(s):
    s = str(s)
    if s.count('%') == 1:
        new_s = s.split('%')
        left_num = abs(float(new_s[0])*100)
        return left_num


def yaml_read_public_config(yaml_name="config.yml"):
    # logger.info("路径：" + os.path.abspath(os.path.dirname(__file__)))
    yml_path = open(yaml_name, encoding="UTF-8")
    # logger.info("路径获取成功！！！")
    data1 = yaml.load(yml_path, Loader=yaml.FullLoader)
    return data1


# 读取yaml文件
def yaml_read_con_per():
    yml_path = 'input.yaml'
    with open(yml_path) as f:
        data1 = yaml.safe_load(f)
    path_pro = data1['FolderList'][0]['Path']
    path_pre = data1['ReferenceFolder']['Path']
    return path_pro, path_pre


if __name__ == '__main__':
    exp_value = 0.1
    file_path = 'D:/Users/20211220/1221pro/compare_result/06_Tower_Foundation_at1220/062_TowerReport_Markov/01_Tower_Markov_Bin128_20211220143456_Confidential/TowerRpt-Markov_Bin128_Markov_Confidential/Tower section=0m Fx _list.csv'
    file_path1 = 'D:/Users/20211220/1221pro/compare_result/06_Tower_Foundation_at1220/062_TowerReport_Markov/01_Tower_Markov_Bin128_20211220143456_Confidential/TowerRpt-Markov_Bin128_Markov_Confidential/Tower section=0m My .csv'

    print(compare_csv_result(file_path1, exp_value))

